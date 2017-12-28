import pickle
import re
import shutil
from difflib import SequenceMatcher
from operator import attrgetter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, NamedStyle

from models import SourceFile, Purchase, Sales, Product, show_progress
from settings import *


def file_exists(file_path):
    return os.path.exists(file_path)


def init():

    print("清理所有错误日志...")
    try:
        shutil.rmtree(os.path.join(ERROR_DIR))
    except FileNotFoundError:
        pass
    os.mkdir(ERROR_DIR)
    print("清理完成.\n")


def clear_output_dirs():
    print("清理所有导出表...")
    try:
        shutil.rmtree(PURCHASE_OUT_DIR)
        shutil.rmtree(SALES_OUT_DIR)
    except FileNotFoundError:
        pass
    years = ['2015', '2016', '2017']
    os.mkdir(PURCHASE_OUT_DIR)
    os.mkdir(SALES_OUT_DIR)
    for y in years:
        os.mkdir(os.path.join(PURCHASE_OUT_DIR, y))
        os.mkdir(os.path.join(SALES_OUT_DIR, y))
    print("清理完成\n")


def save_data(data, file_name, silent=True):
    if not silent:
        print("在文件{}中保存数据...".format(file_name))
    with open(file_name, 'wb') as fout:
        pickle.dump(data, fout, pickle.HIGHEST_PROTOCOL)
    if not silent:
        print("保存完成\n")


def load_data(file_path, silent=True):
    try:
        with open(file_path, "rb") as fin:
            data = pickle.load(fin)
    except FileNotFoundError:
        return None
    if not silent:
        print("从文件{}中读取数据成功!\n".format(file_path))
    return data


def _read_product_info():
    index_sheet = SourceFile(PRODUCT_FILE)
    index_sheet.set_header(0)
    index_sheet.extract_data(model_class=Product)
    return index_sheet.items


def load_products(reload=False):
    print("读取商品目录...")
    if not file_exists(PRODUCT_FILE) or reload:
        products = _read_product_info()
    else:
        products = load_data(PRODUCT_FILE, silent=False)
    print("读取成功!\n")
    return products


def preprocess_sheets():

    def dir_match(path):
        match = re.match(r'.*(?P<new>201\d)', path)
        if not match:
            return False
        return True

    def get_time_info(full_path):
        pattern = r'.*(?P<year>201\d)[\.|年]?(?P<month>\d+).*\.xls'
        matches = re.match(pattern, full_path)
        if matches is not None:
            r = matches.groupdict()
            return r['year'], r['month']
        print("未能成功匹配年月信息")
        return None, None

    print("正在预处理采购/销售数据...")
    purchase_sheets, sales_sheets = [], []
    for dirpath, dirnames, filenames in os.walk(DATA_ROOT):
        if not dir_match(dirpath):
            continue
        for file_name in filenames:
            file_path = os.path.join(dirpath, file_name)
            if '购进' in file_path:
                year, month = get_time_info(file_path)
                sheet = SourceFile(file_path)
                purchase_sheets.append((year, month, sheet))
            elif '销售' in file_path:
                year, month = get_time_info(file_path)
                sheet = SourceFile(file_path)
                sales_sheets.append((year, month, sheet))
    print("预处理完成!\n")
    return purchase_sheets, sales_sheets


def load_ops_list(target='both', reload=False):
    target_texts = {
        "purchase": "采购",
        "sales": "销售",
        "both": "采购/销售"
    }
    print("正在读取{}数据...".format(target_texts[target]))
    if target == 'both':
        p, _ = load_ops_list(target='purchase', reload=reload)
        _, s = load_ops_list(target='sales', reload=reload)
        return p, s

    target_file = PURCHASE_FILE if target == 'purchase' else SALES_FILE
    if not file_exists(target_file) or reload:
        data_list = _generate_ops_list(target)
    else:
        from_file = load_data(target_file, silent=False)
        if target == 'purchase':
            data_list = from_file, []
        else:
            data_list = [], from_file
    print("{}数据读取成功!\n".format(target_texts[target]))
    return data_list


def _generate_ops_list(target='both'):
    p, s = preprocess_sheets()
    purchase_list, sales_list = [], []

    if target == 'both' or target == 'purchase':
        print('开始生成购进数据模型...')
        for year, month, sheet in p:
            sheet.set_header(0)
            sheet.set_time(year, month)
            sheet.extract_data(model_class=Purchase)
            sheet.save_error()
            purchase_list += sheet.items
        print("模型生成完成!\n")

    if target == 'both' or target == 'sales':
        print('开始生成销售数据模型...')
        for year, month, sheet in s:
            sheet.set_header(0)
            sheet.set_time(year, month)
            sheet.extract_data(model_class=Sales)
            sheet.save_error()
            sales_list += sheet.items
        print("模型生成完成!\n")

    return purchase_list, sales_list


def get_string_similarity(s1, s2):
    return SequenceMatcher(None, s1, s2).ratio()


def write_purchase_sheet(vendor_name, year, data):
    if not data:
        return
    wb = Workbook()
    sheet = wb.active
    headers = ["供应商名称", "商品编码", "品名", "规格", "厂家", "数量", '单价', '金额']
    for col in range(len(headers)):
        sheet.cell(row=1, column=col + 1, value=headers[col])
    current_row = 2
    for item in data:
        values = [
            item.vendor, item.serial,
            item.name, item.dose, item.manufacturer,
            item.amount, item.unit_price, item.price
        ]
        if not item.amount or not item.unit_price or not item.price:
            print("时间{}, 第{}行, 信息: {}".format(item.time, item.row, item))
        for i, val in enumerate(values):
            sheet.cell(row=current_row, column=i + 1, value=val)
        current_row += 1
    path = os.path.join(PURCHASE_OUT_DIR, year, "{}.xlsx".format(vendor_name))
    wb.save(path)


def write_sales_sheet(client_name, year, data):
    if not data:
        return
    wb = Workbook()
    sheet = wb.active
    headers = ["客户名称", "商品编码", "品名", "规格", "厂家", "数量", '单价', '金额']
    for col in range(len(headers)):
        sheet.cell(row=1, column=col + 1, value=headers[col])
    current_row = 2
    for item in data:
        values = [
            item.client, item.serial,
            item.name, item.dose, item.manufacturer,
            item.amount, item.unit_price, item.total_price
        ]
        for i, val in enumerate(values):
            sheet.cell(row=current_row, column=i + 1, value=val)
        current_row += 1
    path = os.path.join(SALES_OUT_DIR, year, "{}.xlsx".format(client_name[:20]))
    wb.save(path)


def dump_purchases(vendor_name, purchases):
    data_5, data_6, data_7 = [], [], []
    for i, purchase in enumerate(purchases):
        if purchase.year == '2015':
            data_5.append(purchase)
        elif purchase.year == '2016':
            data_6.append(purchase)
        else:
            data_7.append(purchase)
    write_purchase_sheet(vendor_name, '2015', data_5)
    write_purchase_sheet(vendor_name, '2016', data_6)
    write_purchase_sheet(vendor_name, '2017', data_7)


def dump_sales(client_name, sales):
    data_5, data_6, data_7 = [], [], []
    for i, item in enumerate(sales):
        if item.year == '2015':
            data_5.append(item)
        elif item.year == '2016':
            data_6.append(item)
        else:
            data_7.append(item)
    write_sales_sheet(client_name, '2015', data_5)
    write_sales_sheet(client_name, '2016', data_6)
    write_sales_sheet(client_name, '2017', data_7)


def batch_purchases(purchase_list):
    pl = sorted(purchase_list, key=attrgetter('vendor', 'time'))
    current_vendor = pl[0].vendor
    buff = [pl[0]]
    show_progress(1, len(pl))
    for i, purchase in enumerate(pl, start=1):
        if purchase.vendor == current_vendor:
            buff.append(purchase)
        else:
            dump_purchases(current_vendor, buff)
            current_vendor = purchase.vendor
            buff = [purchase]
        show_progress(i + 1, len(pl))
    dump_purchases(current_vendor, buff)


def batch_sales(sales_list):
    sl = sorted(sales_list, key=attrgetter('client', 'time'))
    current_client = sl[0].client
    buff = [sl[0]]
    for i, sale in enumerate(sl, start=1):
        if sale.client == current_client:
            buff.append(sale)
        else:
            dump_sales(current_client, buff)
            current_client = sale.client
            buff = [sale]
    dump_sales(current_client, buff)


header_style = NamedStyle(
    name='header',
    font=Font(color="ffffff"),
    fill=PatternFill(patternType='solid', fgColor='38761d')
)
error_style = NamedStyle(
    name='error',
    font=Font(color='c93e14', bold=True)
)
purchase_headers = ["类型", "日期", "供应商", "系统编码", "品名", "规格", "生产企业"]
sales_headers = ["类型", "日期", "商品去向", "系统编码", "品名", "规格", "生产企业"]


def get_attribute(item, data_type, attribute):
    if attribute in item.KEY_HEADERS:
        key = item.KEY_HEADERS[attribute]
        return getattr(item, key)
    else:
        if attribute == '类型':
            return "错误项" if data_type != '商品' else "建议"
        if attribute == '日期' and data_type != '商品':
            return "{}.{}".format(item.year, item.month)
        return ""


def write_headers(sheet, headers):
    for col in range(len(headers)):
        cell = sheet.cell(row=1, column=col + 1)
        cell.value = headers[col]
        cell.style = header_style


def write_data_row(item, data_type, sheet, row):
    if data_type == '采购':
        headers = purchase_headers
    else:
        headers = sales_headers

    for col in range(len(headers)):
        value = get_attribute(item, data_type, headers[col])
        cell = sheet.cell(row=row, column=col + 1)
        cell.value = value
        if data_type != '商品':
            cell.style = error_style


def write_match_errors(items, error_type, filename):
    book = Workbook()
    sheet = book.active
    if error_type == '采购':
        write_headers(sheet, purchase_headers)
    else:
        write_headers(sheet, sales_headers)
    current_row = 2
    for item in items:
        write_data_row(item, error_type, sheet, current_row)
        current_row += 1

        def get_similarity(sug):
            a = item.get_product_info
            b = sug.get_product_info
            return get_string_similarity(a, b)
        item.potential_matches.sort(key=get_similarity, reverse=True)

        for suggestion in item.potential_matches[:10]:
            write_data_row(suggestion, '商品', sheet, current_row)
            current_row += 1

    path = os.path.join(ERROR_DIR, filename)
    book.save(path)

if __name__ == '__main__':
    pass
